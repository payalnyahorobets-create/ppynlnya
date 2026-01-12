from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

MONTH_NAMES = [
    "Січень",
    "Лютий",
    "Березень",
    "Квітень",
    "Травень",
    "Червень",
    "Липень",
    "Серпень",
    "Вересень",
    "Жовтень",
    "Листопад",
    "Грудень",
]


def default_excel_path() -> Path:
    return Path(__file__).resolve().parent / "Анализ NEW NOT FINAL (6).xlsx"


@dataclass(frozen=True)
class ExcelStore:
    path: Path

    @lru_cache(maxsize=1)
    def sheet_names(self) -> list[str]:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def month_sheets(self) -> list[str]:
        available = set(self.sheet_names())
        sheets: list[str] = []
        for year in range(2024, 2027):
            for month in MONTH_NAMES:
                name = f"{month} {year}"
                if name in available:
                    sheets.append(name)
        return sheets

    def head_as_records(self, name: str, limit: int = 200) -> list[dict]:
        header, rows = self._read_rows(name, limit)
        records = []
        for row in rows:
            record = {col: value if value is not None else "" for col, value in zip(header, row)}
            records.append(record)
        return records

    def columns(self, name: str) -> list[str]:
        header, _ = self._read_rows(name, 1)
        return [str(col) for col in header]

    def _read_rows(self, name: str, limit: int) -> tuple[list[str], list[tuple]]:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = workbook[name]
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                return [], []
            data_rows: list[tuple] = []
            for index, row in enumerate(rows_iter):
                if index >= limit:
                    break
                data_rows.append(row)
            return [str(col) for col in header], data_rows
        finally:
            workbook.close()

    @lru_cache(maxsize=64)
    def row_count(self, name: str) -> int:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = workbook[name]
            max_row = sheet.max_row or 0
            return max(0, max_row - 1)
        finally:
            workbook.close()


@dataclass(frozen=True)
class ExcelSummary:
    products_count: int
    analysis_count: int
    month_count: int
    month_sheets: list[str]


def build_summary(store: ExcelStore) -> ExcelSummary:
    month_sheets = store.month_sheets()
    return ExcelSummary(
        products_count=store.row_count("Номенклатура"),
        analysis_count=store.row_count("Аналіз ABC-XYZ"),
        month_count=len(month_sheets),
        month_sheets=month_sheets,
    )
